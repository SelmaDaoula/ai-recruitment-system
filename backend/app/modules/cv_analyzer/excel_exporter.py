"""
Module d'export Excel pour les candidats
Génère des fichiers Excel avec données et graphiques
"""

import logging
from typing import List
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models.candidate import Candidate

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exporteur Excel pour les candidats
    """
    
    def __init__(self):
        self.output_dir = Path("data/exports")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info("📊 ExcelExporter initialisé")
    
    def export_candidates(
        self,
        candidates: List[Candidate],
        job_title: str = "Offre d'emploi"
    ) -> str:
        """
        Exporte les candidats en fichier Excel
        
        Args:
            candidates: Liste des candidats
            job_title: Titre de l'offre
        
        Returns:
            str: Chemin du fichier généré
        """
        if not candidates:
            raise ValueError("Aucun candidat à exporter")
        
        logger.info(f"📤 Export de {len(candidates)} candidats...")
        
        # Préparer les données
        data = []
        for c in candidates:
            extracted = c.extracted_data or {}
            
            data.append({
                "ID": c.id,
                "Nom": f"{c.first_name} {c.last_name}",
                "Email": c.email,
                "Téléphone": c.phone,
                "Score CV": c.cv_score,
                "Score Final": c.final_score,
                "Catégorie": self._get_category(c.cv_score),
                "Compétences": len(extracted.get('skills', [])),
                "Expérience (ans)": extracted.get('experience_years', 0),
                "Formation": self._format_education(extracted.get('education', [])),
                "Langues": len(extracted.get('languages', [])),
                "Statut": c.application_status.value if c.application_status else "pending",
                "Date candidature": c.applied_at.strftime("%d/%m/%Y") if c.applied_at else "",
                "Recommandation": c.get_recommendation()
            })
        
        # Créer le DataFrame
        df = pd.DataFrame(data)
        
        # Nom du fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"candidats_{job_title.replace(' ', '_')}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        # Écrire dans Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Feuille 1 : Données
            df.to_excel(writer, sheet_name='Candidats', index=False)
            
            # Feuille 2 : Statistiques
            self._add_statistics_sheet(writer, df)
            
            # Feuille 3 : Détails
            self._add_details_sheet(writer, candidates)
        
        # Ajouter le formatage et les graphiques
        self._format_excel(filepath, df)
        
        logger.info(f"✅ Export terminé : {filename}")
        
        return str(filepath)
    
    def _get_category(self, score: float) -> str:
        """Détermine la catégorie"""
        if score >= 80:
            return "A - Excellent"
        elif score >= 65:
            return "B - Bon"
        elif score >= 50:
            return "C - Moyen"
        else:
            return "D - Insuffisant"
    
    def _format_education(self, education: List) -> str:
        """Formate la formation"""
        if not education:
            return "Non spécifié"
        highest = max(education, key=lambda x: self._education_level(x.get('degree', '')))
        return highest.get('degree', 'Non spécifié')
    
    def _education_level(self, degree: str) -> int:
        """Score du diplôme"""
        degree_lower = degree.lower()
        if 'doctorat' in degree_lower or 'phd' in degree_lower:
            return 5
        elif 'master' in degree_lower or 'ingénieur' in degree_lower:
            return 4
        elif 'licence' in degree_lower or 'bachelor' in degree_lower:
            return 3
        elif 'bts' in degree_lower or 'dut' in degree_lower:
            return 2
        else:
            return 1
    
    def _add_statistics_sheet(self, writer, df):
        """Ajoute une feuille de statistiques"""
        stats_data = {
            "Métrique": [
                "Nombre total de candidats",
                "Score moyen",
                "Score minimum",
                "Score maximum",
                "Catégorie A (≥80)",
                "Catégorie B (65-79)",
                "Catégorie C (50-64)",
                "Catégorie D (<50)"
            ],
            "Valeur": [
                len(df),
                round(df['Score CV'].mean(), 1),
                df['Score CV'].min(),
                df['Score CV'].max(),
                len(df[df['Score CV'] >= 80]),
                len(df[(df['Score CV'] >= 65) & (df['Score CV'] < 80)]),
                len(df[(df['Score CV'] >= 50) & (df['Score CV'] < 65)]),
                len(df[df['Score CV'] < 50])
            ]
        }
        
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
    
    def _add_details_sheet(self, writer, candidates):
        """Ajoute une feuille avec détails complets"""
        details_data = []
        
        for c in candidates:
            extracted = c.extracted_data or {}
            breakdown = c.score_breakdown or {}
            
            details_data.append({
                "Candidat": f"{c.first_name} {c.last_name}",
                "Email": c.email,
                "Téléphone": c.phone,
                "Score Global": c.cv_score,
                "Score Compétences": breakdown.get('skills', 0),
                "Score Expérience": breakdown.get('experience', 0),
                "Score Formation": breakdown.get('education', 0),
                "Score Langues": breakdown.get('languages', 0),
                "Compétences": ", ".join(extracted.get('skills', [])[:10]),
                "Expérience": f"{extracted.get('experience_years', 0)} ans",
                "Recommandation": c.get_recommendation()
            })
        
        details_df = pd.DataFrame(details_data)
        details_df.to_excel(writer, sheet_name='Détails', index=False)
    
    def _format_excel(self, filepath, df):
        """Ajoute le formatage et les graphiques"""
        wb = load_workbook(filepath)
        
        # Formater la feuille Candidats
        ws = wb['Candidats']
        
        # En-têtes en gras avec fond bleu
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Colorer les scores
        for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=5, max_col=5):
            for cell in row:
                if cell.value:
                    if cell.value >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif cell.value >= 65:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif cell.value < 50:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Ajouter un graphique dans Statistiques
        self._add_chart(wb)
        
        wb.save(filepath)
    
    def _add_chart(self, wb):
        """Ajoute un graphique de distribution"""
        ws = wb['Statistiques']
        
        # Graphique en barres pour les catégories
        chart = BarChart()
        chart.title = "Distribution des candidats par catégorie"
        chart.x_axis.title = "Catégorie"
        chart.y_axis.title = "Nombre de candidats"
        
        data = Reference(ws, min_col=2, min_row=5, max_row=8)
        cats = Reference(ws, min_col=1, min_row=5, max_row=8)
        
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D5")


# ============ Test de l'exporteur ============

if __name__ == "__main__":
    print("\n" + "="*60)
    print("MODULE D'EXPORT EXCEL")
    print("="*60)
    print("\nCe module permet d'exporter les candidats en Excel avec :")
    print("  • Données complètes des candidats")
    print("  • Statistiques détaillées")
    print("  • Graphiques de distribution")
    print("  • Formatage professionnel")
    print("  • Couleurs selon les scores")
    print("\n" + "="*60 + "\n")